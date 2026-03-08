"""
export_to_site.py
=================
Reads your WebsiteData.xlsx workbook (4 tabs) and exports JSON for the website.

SETUP (one time):
    pip install openpyxl

USAGE:
    1. Update WORKBOOK_PATH below to match your file location
    2. Run: python export_to_site.py
    3. JSON files are saved to the public/data folder

YOUR WORKBOOK TABS:
    - Projections  → projections.json
    - Record       → record.json
    - Rankings     → rankings.json
    - MMSim        → bracket.json
"""

import json
import os
import subprocess
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    subprocess.run(["pip", "install", "openpyxl"], check=True)
    import openpyxl


# ============================================================
# CONFIGURATION — Update this path
# ============================================================

WORKBOOK_PATH = "WebsiteData.xlsx"  # <-- Your single Excel file

OUTPUT_DIR = "public/data"

# GitHub repo path (set to your repo folder to enable auto-push, or None to skip)
GITHUB_REPO_DIR = None


# ============================================================
# COLUMN MAPPINGS — Matched to your spreadsheet headers
# ============================================================

# Projections tab:
#   Game | Team | TeamOpp | Margin | Blowout Potential | Kelly % | EV % | Bet | Units
#
# Data formats:
#   Margin:           raw number (17.9 = Team favored by 17.9)
#   Blowout Potential: decimal   (0.798 = 79.8%)
#   Kelly %:          decimal    (0.077 = 7.7%)
#   EV %:             decimal    (0.244 = 24.4%)
#   Bet:              "Yes" or "No"
#   Units:            raw number (1.5 = 1.5 units)

PROJ_GAME = "Game"
PROJ_TEAM1 = "Team"
PROJ_TEAM2 = "TeamOpp"
PROJ_POSITION = "Team Position"
PROJ_MARGIN = "Margin"
PROJ_KELLY = "Kelly %"
PROJ_EV = "EV %"
PROJ_BLOWOUT = "Blowout Potential"
PROJ_MARKET = "Market Spread"
PROJ_BET = "Bet"
PROJ_UNITS = "Units"

# Rankings tab:
#   Team | Rating | New Rank
RANK_TEAM = "Team"
RANK_RATING = "Rating"
RANK_RANK = "New Rank"
RANK_LAST = "Last Rank"

# MMSim tab:
#   Team | Round of 32 Prob | Round of 16 Prob | Elite 8 Prob |
#   Final Four Prob | Champ Game Prob | Champion Prob | Region | Seed | Rating
SIM_TEAM = "Team"
SIM_R32 = "Round of 32 Prob"
SIM_R16 = "Round of 16 Prob"
SIM_E8 = "Elite 8 Prob"
SIM_F4 = "Final Four Prob"
SIM_CG = "Champ Game Prob"
SIM_CHAMP = "Champion Prob"
SIM_REGION = "Region"
SIM_SEED = "Seed"
SIM_RATING = "Rating"


# ============================================================
# HELPERS
# ============================================================

def read_sheet(wb, sheet_name):
    if sheet_name not in wb.sheetnames:
        print(f"   ⚠️  Tab '{sheet_name}' not found. Available: {wb.sheetnames}")
        return None
    ws = wb[sheet_name]
    headers = [str(c.value).strip() if c.value else f"col_{c.column}"
               for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if all(v is None for v in row):
            continue
        rows.append(dict(zip(headers, row)))
    return rows


def col(row, name):
    for k, v in row.items():
        if k.lower().strip() == name.lower().strip():
            return v
    return None


def to_float(val, default=0.0):
    if val is None: return default
    try: return round(float(val), 6)
    except: return default


def to_int(val, default=0):
    if val is None: return default
    try: return int(float(val))
    except: return default


# ============================================================
# EXPORTERS
# ============================================================

def export_projections(wb):
    rows = read_sheet(wb, "Projections")
    if not rows: return None
    games = []
    for i, r in enumerate(rows):
        t1, t2 = col(r, PROJ_TEAM1), col(r, PROJ_TEAM2)
        if not t1 or not t2: continue

        margin = to_float(col(r, PROJ_MARGIN))
        market = to_float(col(r, PROJ_MARKET))
        kelly_raw = to_float(col(r, PROJ_KELLY))       # decimal → %
        ev_raw = to_float(col(r, PROJ_EV))              # decimal → %
        blowout_raw = to_float(col(r, PROJ_BLOWOUT))    # decimal → %
        units_raw = to_float(col(r, PROJ_UNITS))         # raw number
        bet_raw = col(r, PROJ_BET)                        # "Yes" or "No"
        is_bet = str(bet_raw).strip().lower() == "yes" if bet_raw else False
        position_raw = col(r, PROJ_POSITION)
        position = str(position_raw).strip() if position_raw else ""

        games.append({
            "id": i + 1,
            "team1": str(t1).strip(),
            "team2": str(t2).strip(),
            "position": position,
            "spread": round(margin, 1),
	    "market": round(market, 1),
            "kelly": round(kelly_raw * 100, 1),
            "ev": round(ev_raw * 100, 1),
            "blowout": round(blowout_raw * 100),
            "units": round(units_raw, 3),
            "bet": is_bet,
        })

    print(f"   ✅ Projections: {len(games)} games")
    return {
        "date": datetime.now().strftime("%b %d, %Y"),
        "updated": datetime.now().strftime("%I:%M %p ET"),
        "games": games,
    }


def export_rankings(wb):
    rows = read_sheet(wb, "Rankings")
    if not rows: return None
    teams = []
    for r in rows:
        team = col(r, RANK_TEAM)
        if not team: continue
        teams.append({
           "rank": to_int(col(r, RANK_RANK)),
            "lastRank": to_int(col(r, RANK_LAST)),
            "team": str(team).strip(),
            "rating": round(to_float(col(r, RANK_RATING)), 3),
        })
    teams.sort(key=lambda t: t["rank"])
    print(f"   ✅ Rankings: {len(teams)} teams")
    return {
        "week": datetime.now().strftime("Week of %b %d, %Y"),
        "updated": datetime.now().strftime("%b %d, %Y"),
        "teams": teams,
    }


def export_bracket(wb):
    rows = read_sheet(wb, "MMSim")
    if not rows: return None
    regions, champ_odds = {}, []
    for r in rows:
        team, region = col(r, SIM_TEAM), col(r, SIM_REGION)
        if not team or not region: continue
        team, region = str(team).strip(), str(region).strip()
        seed = to_int(col(r, SIM_SEED))
        td = {
            "seed": seed, "team": team,
            "rating": round(to_float(col(r, SIM_RATING)), 1),
            "r32": round(to_float(col(r, SIM_R32)) * 100, 1),
            "r16": round(to_float(col(r, SIM_R16)) * 100, 1),
            "e8": round(to_float(col(r, SIM_E8)) * 100, 1),
            "f4": round(to_float(col(r, SIM_F4)) * 100, 1),
            "cg": round(to_float(col(r, SIM_CG)) * 100, 1),
            "champ": round(to_float(col(r, SIM_CHAMP)) * 100, 1),
        }
        regions.setdefault(region, []).append(td)
        if td["champ"] >= 0.1:
            champ_odds.append({"seed": seed, "team": team, "region": region, "champ": td["champ"]})
    for reg in regions:
        regions[reg].sort(key=lambda t: t["seed"])
    champ_odds.sort(key=lambda t: t["champ"], reverse=True)
    total = sum(len(t) for t in regions.values())
    print(f"   ✅ Bracket: {total} teams, {len(regions)} regions, {len(champ_odds)} with title odds")
    return {"updated": datetime.now().strftime("%b %d, %Y"), "regions": regions, "champOdds": champ_odds}


def export_record(wb):
    """
    Export Record tab to record.json

    Row 1: Headers — W, L, ROI, CLV, Last 10W, Last10L
    Row 2: Summary values
    Row 5: Headers — Date, Game #, Win/Loss/Push, Stake, Profit, Running Profit, CLV, Running CLV
    Row 6+: Daily history
    """
    if "Record" not in wb.sheetnames:
        print(f"   ⚠️  Tab 'Record' not found.")
        return None

    ws = wb["Record"]
    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))

    # Row 1 = headers, Row 2 = summary values
    summary = {}
    if len(rows) >= 2:
        headers_top = [str(h).strip() if h else "" for h in rows[0]]
        values_top = rows[1]
        for h, v in zip(headers_top, values_top):
            if h:
                summary[h] = v

    record = {
        "w": to_int(summary.get("W")),
        "l": to_int(summary.get("L")),
        "roi": to_float(summary.get("ROI")),
        "clv": to_float(summary.get("CLV")),
        "last10w": to_int(summary.get("Last 10W", summary.get("Last10W"))),
        "last10l": to_int(summary.get("Last10L", summary.get("Last 10L"))),
    }

    # Row 5+ = daily history
    history = []
    if len(rows) >= 5:
        hist_headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[4])]
        for row in rows[5:]:
            if all(v is None for v in row):
                continue
            entry = dict(zip(hist_headers, row))
            date_val = entry.get("Date")
            running_profit = entry.get("Running Profit")
            running_clv = entry.get("Running CLV")
            result = entry.get("Win/Loss/Push")

            if date_val is None and running_profit is None:
                continue

            # Format date
            date_str = ""
            if date_val:
                if hasattr(date_val, 'strftime'):
                    date_str = date_val.strftime("%b %d")
                else:
                    date_str = str(date_val)

            h = {
                "date": date_str,
                "result": str(result).strip() if result else None,
                "stake": to_float(entry.get("Stake")),
                "profit": to_float(entry.get("Profit")),
                "runningProfit": to_float(running_profit),
                "clv": to_float(entry.get("CLV")),
                "runningCLV": to_float(running_clv),
            }
            history.append(h)

    print(f"   ✅ Record: {record['w']}-{record['l']}, {len(history)} history entries")
    return {
        "summary": record,
        "history": history,
    }


# ============================================================
# SAVE & PUSH
# ============================================================

def save_json(data, filename):
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    path = os.path.join(OUTPUT_DIR, filename)
    with open(path, "w") as f:
        json.dump(data, f, indent=2)
    print(f"   💾 {path}")


def push_to_github():
    if not GITHUB_REPO_DIR:
        print(f"\n📤 To enable auto-push, set GITHUB_REPO_DIR in this script.")
        return
    try:
        os.chdir(GITHUB_REPO_DIR)
        subprocess.run(["git", "add", f"{OUTPUT_DIR}/"], check=True)
        msg = f"Update data — {datetime.now().strftime('%b %d %Y %I:%M %p')}"
        subprocess.run(["git", "commit", "-m", msg], check=True)
        subprocess.run(["git", "push"], check=True)
        print("   ✅ Pushed to GitHub!")
    except subprocess.CalledProcessError as e:
        print(f"   ⚠️  Git push failed: {e}")


# ============================================================
# MAIN
# ============================================================

def main():
    print("=" * 50)
    print("🏀 Jack's CBB — Data Export")
    print("=" * 50)
    print(f"\n📂 Reading: {WORKBOOK_PATH}")

    if not os.path.exists(WORKBOOK_PATH):
        print(f"\n⚠️  File not found: {WORKBOOK_PATH}")
        print(f"   Update WORKBOOK_PATH at the top of this script.")
        return

    wb = openpyxl.load_workbook(WORKBOOK_PATH, data_only=True)
    print(f"   Tabs found: {wb.sheetnames}")

    projections = export_projections(wb)
    rankings = export_rankings(wb)
    bracket = export_bracket(wb)
    record = export_record(wb)

    print(f"\n💾 Saving...")
    if projections: save_json(projections, "projections.json")
    if rankings:    save_json(rankings, "rankings.json")
    if bracket:     save_json(bracket, "bracket.json")
    if record:      save_json(record, "record.json")

    push_to_github()

    print(f"\n{'=' * 50}")
    print(f"✅ Done! Files in ./{OUTPUT_DIR}/")
    print(f"{'=' * 50}")


if __name__ == "__main__":
    main()
